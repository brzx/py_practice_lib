import time
import numbers
import datetime
import warnings
import openpyxl
from collections import namedtuple
from collections import OrderedDict
from openpyxl.styles import PatternFill
from functools import reduce
import sys, pdb

DefaultStartCellValue = 'ID'
DefaultColumnStart = 'A'
DefaultRowStart = 13

def processFileterValue(value):
    ''' now we can process =, >, >=, <, <= '''
    if len(value) > 1:
        if value[1] == '=':
            try:
                vl = float(value[2:])
            except ValueError:
                vl = value[2:]
            rl = [value[:2], vl, ]
        else:
            try:
                vl = float(value[1:])
            except ValueError:
                vl = value[1:]
            rl = ['==' if value[0]=='=' else value[0], vl, ]
    else:
        print('Filter value is not correct.')
        sys.exit()
    return rl

def getConf():
    ''' get configuration data from an excel file '''
    try:
        wbconf = openpyxl.load_workbook('configuration.xlsx')
    except FileNotFoundError as err:
        print(err)
        sys.exit()
    ws = wbconf['comparison']
    collist = [cel.value.strip() for cel in ws['A']][1:]
    confcoldict = OrderedDict({cv:idx for idx, cv in enumerate(collist, start=1)})
    splitcell = filter(lambda x: x.value is not None and x.value.upper()=='Y',
                    [cel for cel in ws['B']][1:])
    splitlist = tuple([str(ws.cell(row=cel.row, column=cel.col_idx-1).value)
                    for cel in splitcell])
    filtercell = filter(lambda x: x.value is not None and x.value.upper()=='Y',
                    [cel for cel in ws['C']][1:])
    filterdict = {str(ws.cell(row=cel.row, column=cel.col_idx-2).value).strip():
                    processFileterValue(ws.cell(row=cel.row, 
                        column=cel.col_idx+1).value)
                    for cel in filtercell}
    confcoldict = {cv:idx for idx, cv in enumerate(collist, start=1)}
    try:
        oldFile = ws['F2'].value.strip()
        oldSheetName = ws['G2'].value.strip()
        oldRowStart = ws['H2'].value
        oldColStart = ws['I2'].value
        newFile = ws['F3'].value.strip()
        newSheetName = ws['G3'].value.strip()
        newRowStart = ws['H3'].value
        newColStart = ws['I3'].value
        secqformat = ws['G5'].value.strip()
        sec1oldq = [secqformat.format(vl) for vl in list(str(ws['F5'].value))]
        sec1newq = [secqformat.format(vl) for vl in list(str(ws['F6'].value))]
        sec2oldq = [secqformat.format(vl) for vl in list(str(ws['F7'].value))]
        sec2newq = [secqformat.format(vl) for vl in list(str(ws['F8'].value))]
        sec3oldq = [secqformat.format(vl) for vl in list(str(ws['F9'].value))]
        sec3newq = [secqformat.format(vl) for vl in list(str(ws['F10'].value))]
        sec4oldq = [secqformat.format(vl) for vl in list(str(ws['F11'].value))]
        sec4newq = [secqformat.format(vl) for vl in list(str(ws['F12'].value))]
    except Exception as err:
        print(err)
        sys.exit()
    wbconf.close()
    return {
        'confcoldict': confcoldict,
        'splitlist': splitlist,
        'filterdict': filterdict,
        'oldFile': oldFile,
        'oldSheetName': oldSheetName,
        'oldRowStart': oldRowStart,
        'oldColStart': oldColStart,
        'newFile': newFile,
        'newSheetName': newSheetName,
        'newRowStart': newRowStart,
        'newColStart': newColStart,
        'sec1oldq': sec1oldq,
        'sec1newq': sec1newq,
        'sec2oldq': sec2oldq,
        'sec2newq': sec2newq,
        'sec3oldq': sec3oldq,
        'sec3newq': sec3newq,
        'sec4oldq': sec4oldq,
        'sec4newq': sec4newq,
        'secqformat': secqformat,
    }

def printConf(conf):
    ''' print configuration information '''
    print('------------------------------------ configuration start ------------'
          '------------------------------\n')
    for key, value in conf.items():
        print('{0:15} : {1}'.format(key, value))
    print('------------------------------------ configuration end --------------'
          '------------------------------\n')

DataRange = namedtuple(
            'DataRange', 
            ['Col'+str(i) for i in getConf().get('confcoldict').values()]
        )

def makeRangeDict(filename, 
                  sheet, 
                  rowstart=DefaultRowStart, 
                  colstart=DefaultColumnStart):
    ''' make orginal data range with column name and row end number 
        from the source sheet
    '''
    if sheet[DefaultColumnStart+str(rowstart-1)].value.strip() \
            == DefaultStartCellValue:
        coldict = {}
        for idy, cd in enumerate(sheet[rowstart-1], start=1):
            if cd.value is not None:
                coldict[cd.value.strip()] = idy
            else:
                break
        return {'rowend': sheet[DefaultColumnStart][-1].row, 'coldict': coldict, }
    else:
        warnings.warn('File {} format is not correct.'.format(filename))
        return None

def getColumnHeader(conf):
    ''' get target column name as header '''
    return DataRange(*conf.get('confcoldict').keys())

def getDR(id, drlist):
    ''' get DR by ID '''
    return next(filter(lambda x: x.Col1==id, drlist))

def checkColumns(conf, rangeDict):
    ''' check whether all the provided columns 
        are in source and target files
    '''
    for col in conf.get('confcoldict').keys():
        if col not in rangeDict.keys():
            print('Column {} not found!'.format(col))
            sys.exit()

def getDataBlock(conf, filename, sheetname, 
                 rowstart=DefaultRowStart,
                 colstart=DefaultColumnStart,
                 full=False):
    ''' get the data block from provided file '''
    try:
        inFile = openpyxl.load_workbook(conf.get(filename))
    except FileNotFoundError as err:
        print(err)
        sys.exit()
    try:
        inFileSheet = inFile[conf.get(sheetname)]
    except KeyError as err:
        print(err, end=' ')
        print('in the {}'.format(filename))
        sys.exit()
    inrd = makeRangeDict(conf.get(filename), inFileSheet, rowstart, colstart)
    checkColumns(conf, inrd.get('coldict'))
    
    if inrd is not None:
        outlist = []
        rowend = inrd.get('rowend')
        mapdict = {}
        for colx, nux in conf.get('confcoldict').items():
            for coly, nuy in inrd.get('coldict').items():
                if colx == coly:
                    mapdict['Col'+str(nux)] = nuy
        while rowstart <= rowend:
            tempdict = dict(mapdict)
            for col, cv in tempdict.items():
                tempvalue = inFileSheet.cell(row=rowstart, column=cv).value
                if isinstance(tempvalue, str):
                    tempdict[col] = tempvalue.strip()
                else:
                    tempdict[col] = tempvalue
            try:
                dr = DataRange(**tempdict)
            except TypeError as te:
                print(te)
            else:
                if full is not True:
                    # do the filter operation
                    flag = True
                    for fkey, fvalue in conf.get('filterdict').items():
                        if isinstance(fvalue[1], numbers.Real):
                            vl = getattr(dr, 'Col'+str(conf.get('confcoldict').get(fkey)))
                            evalstr = '{0} {1} {2}'.format(
                                vl if vl is not None else 0,
                                fvalue[0],
                                fvalue[1]
                            )
                        else:
                            vl = getattr(dr, 'Col'+str(conf.get('confcoldict').get(fkey)))
                            evalstr = "'{0}' {1} '{2}'".format(
                                vl if vl is not None else '',
                                fvalue[0],
                                fvalue[1]
                            )
                        
                        if not eval(evalstr):
                            flag = False
                    if flag:
                        outlist.append(dr)
                else:
                    outlist.append(dr)
            del tempdict
            rowstart += 1
    else:
        print('No data found!')
    inFile.close()
    return outlist

def getSectionSet(conf, dataBlock, secflag):
    return {
        odb.Col1
        for odb in list(filter(
            lambda x: 
                getattr(x, 'Col'+str(conf.get('confcoldict').get('ECSD Quarter'))) \
                    in conf.get(secflag),
            dataBlock
        ))
    }

def compare(conf):
    ''' compare old data with new data, return not changed 
        and changed data block
    '''
    oldDataBlock = getDataBlock(conf, 
                                'oldFile', 
                                'oldSheetName', 
                                rowstart=conf.get('oldRowStart'),
                                colstart=conf.get('oldColStart'))
    newDataBlock = getDataBlock(conf, 
                                'newFile', 
                                'newSheetName', 
                                rowstart=conf.get('newRowStart'),
                                colstart=conf.get('newColStart'))

    oldset4Sec1 = getSectionSet(conf, oldDataBlock, 'sec1oldq')
    newset4Sec1 = getSectionSet(conf, newDataBlock, 'sec1newq')
    res4Sec1 = sorted(list(oldset4Sec1-newset4Sec1))
    print('res4Sec1 is {}'.format(res4Sec1))
    if len(res4Sec1) > 0:
        block4Sec1 = [db for db in oldDataBlock if db.Col1 in res4Sec1]
    else:
        block4Sec1 = []
    
    oldset4Sec2 = getSectionSet(conf, oldDataBlock, 'sec2oldq')
    newset4Sec2 = getSectionSet(conf, newDataBlock, 'sec2newq')
    res4Sec2 = sorted(list(oldset4Sec2&newset4Sec2))
    print('res4Sec2 is {}'.format(res4Sec2))
    if len(res4Sec2) > 0:
        block4Sec2 = [db for db in newDataBlock if db.Col1 in res4Sec2]
    else:
        block4Sec2 = []
    
    oldset4Sec3 = getSectionSet(conf, oldDataBlock, 'sec3oldq')
    newset4Sec3 = getSectionSet(conf, newDataBlock, 'sec3newq')
    res4Sec3 = sorted(list(newset4Sec3-oldset4Sec3))
    print('res4Sec3 is {}'.format(res4Sec3))
    if len(res4Sec3) > 0:
        block4Sec3 = [db for db in newDataBlock if db.Col1 in res4Sec3]
    else:
        block4Sec3 = []
    
    oldset4Sec4 = getSectionSet(conf, oldDataBlock, 'sec4oldq')
    newset4Sec4 = getSectionSet(conf, newDataBlock, 'sec4newq')
    res4Sec4 = sorted(list(oldset4Sec4&newset4Sec4))
    print('res4Sec4 is {}'.format(res4Sec4))
    block4Sec4 = []
    if len(res4Sec4) > 0:
        for it in res4Sec4:
            try:
                olddr = getDR(it, oldDataBlock)
                newdr = getDR(it, newDataBlock)
            except StopIteration as err:
                print(err)
            else:
                changelist = {'changed': 'N'}
                for sl in conf.get('splitlist'):
                    if getattr(olddr, 'Col'+str(conf.get('confcoldict').get(sl))) != \
                        getattr(newdr, 'Col'+str(conf.get('confcoldict').get(sl))):
                        changelist[sl] = 'Y'
                        changelist['changed'] = 'Y'
                    else:
                        changelist[sl] = 'N'
                block4Sec4.append((olddr, newdr, changelist))
    
    def printComapre(db, nm):
        print('----------------------------- {}, length: {} ---------------'
                '---------------------\n'.format(nm, len(db)))
        for vl in db:
            print(vl)
    print('------------------------------------ compare start ------------------'
              '------------------------------\n')
    printComapre(block4Sec1, 'block4Sec1')
    printComapre(block4Sec2, 'block4Sec2')
    printComapre(block4Sec3, 'block4Sec3')
    printComapre(block4Sec4, 'block4Sec4')
    print('------------------------------------ compare end --------------------'
              '------------------------------\n')
    return {
        'block4Sec1': block4Sec1,
        'block4Sec2': block4Sec2,
        'block4Sec3': block4Sec3,
        'block4Sec4': block4Sec4,
    }

def writeHeader(conf, ws, row):
    ''' write final column header to the sheet '''
    idx = 1
    for colnm in conf.get('confcoldict').keys():
        if colnm in conf.get('splitlist'):
            ws.cell(row=row, column=idx, value=colnm+' old')
            idx += 1
            ws.cell(row=row, column=idx, value=colnm+' new')
            idx += 1
        else:
            ws.cell(row=row, column=idx, value=colnm)
            idx += 1
    return OrderedDict({cel.value:cel.col_idx for cel in ws[row] 
            if cel.value is not None})

def writeRowBackground(conf, ws, row, color):
    fill = PatternFill("solid", fgColor=color)
    for colnb in range(len(conf.get('confcoldict'))+len(conf.get('splitlist'))):
        ws.cell(row=row, column=colnb+1).fill = fill

def writeColBackground(conf, ws, row, headermap, color):
    fill = PatternFill("solid", fgColor=color)
    newnames = [cn+' old' for cn in conf.get('splitlist')] + \
               [cn+' new' for cn in conf.get('splitlist')]
    for colnb in [nb for cn, nb in headermap.items() if cn in newnames]:
        ws.cell(row=row, column=colnb).fill = fill

def writeBlock(conf, dataBlock):
    ''' write compared data to a new xlsx file '''
    filename = 'PRD GC Technology Landing - Compared_' \
        + datetime.datetime.now().strftime('%Y%m%d') \
        + '_' + str(time.time()) + '.xlsx'
    wbTarget = openpyxl.Workbook()
    ws = wbTarget.active

    # start to write section one
    rowgo = 1
    ws.cell(row=rowgo, column=1, 
            value='Section 1: Data in old but not in new start')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')
    rowgo += 1
    headermap = writeHeader(conf, ws, rowgo)
    writeRowBackground(conf, ws, rowgo, '5CACEE')
    writeColBackground(conf, ws, rowgo, headermap, '548B54')
    rowgo += 1
    if len(dataBlock.get('block4Sec1')) > 0:
        for dr in dataBlock.get('block4Sec1'):
            for nm, nb in conf.get('confcoldict').items():
                if nm in conf.get('splitlist'):
                    ws.cell(row=rowgo, column=headermap.get(nm+' old'),
                            value=getattr(dr, 'Col'+str(nb)))
                else:
                    ws.cell(row=rowgo, column=headermap.get(nm),
                            value=getattr(dr, 'Col'+str(nb)))
            rowgo += 1
    ws.cell(row=rowgo, column=1, 
            value='Section 1: Data in old but not in new end')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')
    rowgo += 3

    # start to write section two
    ws.cell(row=rowgo, column=1, 
            value='Section 2: slip to following quarter start')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')
    rowgo += 1
    headermap = writeHeader(conf, ws, rowgo)
    writeRowBackground(conf, ws, rowgo, '5CACEE')
    writeColBackground(conf, ws, rowgo, headermap, '548B54')
    rowgo += 1
    if len(dataBlock.get('block4Sec2')) > 0:
        for dr in dataBlock.get('block4Sec2'):
            for nm, nb in conf.get('confcoldict').items():
                if nm in conf.get('splitlist'):
                    ws.cell(row=rowgo, column=headermap.get(nm+' new'),
                            value=getattr(dr, 'Col'+str(nb)))
                else:
                    ws.cell(row=rowgo, column=headermap.get(nm),
                            value=getattr(dr, 'Col'+str(nb)))
            rowgo += 1
    ws.cell(row=rowgo, column=1, 
            value='Section 2: slip to following quarter end')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')
    rowgo += 3

    # start to write section three
    ws.cell(row=rowgo, column=1, 
            value='Section 3: Data in new but not in old start')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')
    rowgo += 1
    headermap = writeHeader(conf, ws, rowgo)
    writeRowBackground(conf, ws, rowgo, '5CACEE')
    writeColBackground(conf, ws, rowgo, headermap, '548B54')
    rowgo += 1
    if len(dataBlock.get('block4Sec3')) > 0:
        for dr in dataBlock.get('block4Sec3'):
            for nm, nb in conf.get('confcoldict').items():
                if nm in conf.get('splitlist'):
                    ws.cell(row=rowgo, column=headermap.get(nm+' new'),
                            value=getattr(dr, 'Col'+str(nb)))
                else:
                    ws.cell(row=rowgo, column=headermap.get(nm),
                            value=getattr(dr, 'Col'+str(nb)))
            rowgo += 1
    ws.cell(row=rowgo, column=1, 
            value='Section 3: Data in new but not in old end')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')
    rowgo += 3
    
    # start to write section four
    ws.cell(row=rowgo, column=1, 
            value='Section 4: Data in both new and old start')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')
    rowgo += 1
    headermap = writeHeader(conf, ws, rowgo)
    writeRowBackground(conf, ws, rowgo, '5CACEE')
    writeColBackground(conf, ws, rowgo, headermap, '548B54')
    rowgo += 1
    if len(dataBlock.get('block4Sec4')) > 0:
        for drold, drnew, flag in dataBlock.get('block4Sec4'):
            for nm, nb in conf.get('confcoldict').items():
                if nm in conf.get('splitlist'):
                    if flag.get(nm) == 'Y':
                        ws.cell(row=rowgo, column=headermap.get(nm+' new'),
                                value=getattr(drnew, 'Col'+str(nb)))
                        ws.cell(row=rowgo, column=headermap.get(nm+' old'),
                                value=getattr(drold, 'Col'+str(nb)))
                    else:
                        ws.cell(row=rowgo, column=headermap.get(nm+' new'),
                                value=getattr(drnew, 'Col'+str(nb)))
                    if flag.get('changed') == 'Y':
                        writeRowBackground(conf, ws, rowgo, 'FFEC8B')
                else:
                    ws.cell(row=rowgo, column=headermap.get(nm),
                            value=getattr(drnew, 'Col'+str(nb)))
            rowgo += 1
    ws.cell(row=rowgo, column=1, 
            value='Section 4: Data in both new and old end')
    writeRowBackground(conf, ws, rowgo, 'FFA07A')

    wbTarget.save(filename)
    wbTarget.close()
    return filename

def checkDataBlock(conf, datablock):
    propertyList = ['Col'+str(nb) for nb in range(1, len(conf.get('confcoldict'))+1)]
    for idx, dr in enumerate(datablock):
        try:
            for pl in propertyList:
                getattr(dr, pl)
        except Exception as err:
            print('index: {}, Col1: {}.'.format(idx, getattr(dr, 'Col1')))
            print(err)

def newReport(conf, filename):
    fullnewDataBlockAll = getDataBlock(conf, 
        'newFile', 
        'newSheetName', 
        rowstart=conf.get('newRowStart'),
        colstart=conf.get('newColStart'),
        full=True)
    fullnewDataBlock = list(filter(
        lambda x: getattr(x, 'Col1') is not None,
        fullnewDataBlockAll
    ))
    checkDataBlock(conf, fullnewDataBlock)
    
    def printReportData(db, name):
        print('------------------------------------ %s start ------------------'
                  '------------------------------\n' % name)
        for fn in db:
            print(fn)
        print('------------------------------------ %s end --------------------'
                  '------------------------------\n' % name)

    ColReportingStatus = 'Col' + str(conf.get('confcoldict').get('Reporting Status'))
    ColECSDQuarter = 'Col' + str(conf.get('confcoldict').get('ECSD Quarter'))
    ColASNetRevenue = 'Col' + str(conf.get('confcoldict').get('AS Net Revenue'))
    ColWtdASNetRevenue = 'Col' + str(conf.get('confcoldict').get('Wtd AS Net Revenue'))
    ColSINetRevenue = 'Col' + str(conf.get('confcoldict').get('SI Net Revenue'))
    ColWtdSINetRevenue = 'Col' + str(conf.get('confcoldict').get('Wtd SI Net Revenue'))
    ColAONetRevenue = 'Col' + str(conf.get('confcoldict').get('AO Net Revenue'))
    ColWtdAONetRevenue = 'Col' + str(conf.get('confcoldict').get('Wtd AO Net Revenue'))

    def sumValue(dblist, column):
        sumlist = [
            float(getattr(dv, column)) 
            for dv in dblist 
            if isinstance(getattr(dv, column), numbers.Real)
        ]
        return sum(sumlist)

    def getSumDict(dblist):
        sumDict = {}
        if len(dblist) > 1:
            try:
                sumDict['ASNetRevenue'] = sumValue(dblist, ColASNetRevenue)
                sumDict['WtdASNetRevenue'] = sumValue(dblist, ColWtdASNetRevenue)
                sumDict['SINetRevenue'] = sumValue(dblist, ColSINetRevenue)
                sumDict['WtdSINetRevenue'] = sumValue(dblist, ColWtdSINetRevenue)
                sumDict['AONetRevenue'] = sumValue(dblist, ColAONetRevenue)
                sumDict['WtdAONetRevenue'] = sumValue(dblist, ColWtdAONetRevenue)
            except ValueError as err:
                print(err)
                print('There should be some value which is not number type.')
                sys.exit()

        elif len(dblist) == 1:
            try:
                sumDict['ASNetRevenue'] = float(getattr(dblist[0], ColASNetRevenue))
                sumDict['WtdASNetRevenue'] = float(getattr(dblist[0], ColWtdASNetRevenue))
                sumDict['SINetRevenue'] = float(getattr(dblist[0], ColSINetRevenue))
                sumDict['WtdSINetRevenue'] = float(getattr(dblist[0], ColWtdSINetRevenue))
                sumDict['AONetRevenue'] = float(getattr(dblist[0], ColAONetRevenue))
                sumDict['WtdAONetRevenue'] = float(getattr(dblist[0], ColWtdAONetRevenue))
            except ValueError as err:
                print(err)
                print('There should be some value which is not number type.')
                sys.exit()
        else:
            pass
        return sumDict

    Q2Qualified = list(filter(
        lambda x: getattr(x, ColReportingStatus)=='Pipeline' \
            and getattr(x, ColECSDQuarter)=='FY19Q2',
        fullnewDataBlock
    ))
    printReportData(Q2Qualified, 'Q2Qualified')
    Q2QualifiedSumDict = getSumDict(Q2Qualified)
    print(Q2QualifiedSumDict)

    Q2Unqualified = list(filter(
        lambda x: getattr(x, ColReportingStatus)=='Unqualified' \
            and getattr(x, ColECSDQuarter)=='FY19Q2',
        fullnewDataBlock
    ))
    printReportData(Q2Unqualified, 'Q2Unqualified')
    Q2UnqualifiedSumDict = getSumDict(Q2Unqualified)
    print(Q2UnqualifiedSumDict)

    Q3Qualified = list(filter(
        lambda x: getattr(x, ColReportingStatus)=='Pipeline' \
            and getattr(x, ColECSDQuarter)=='FY19Q3',
        fullnewDataBlock
    ))
    printReportData(Q3Qualified, 'Q3Qualified')
    Q3QualifiedSumDict = getSumDict(Q3Qualified)
    print(Q3QualifiedSumDict)

    Q3Unqualified = list(filter(
        lambda x: getattr(x, ColReportingStatus)=='Unqualified' \
            and getattr(x, ColECSDQuarter)=='FY19Q3',
        fullnewDataBlock
    ))
    printReportData(Q3Unqualified, 'Q3Unqualified')
    Q3UnqualifiedSumDict = getSumDict(Q3Unqualified)
    print(Q3UnqualifiedSumDict)

    Q4Qualified = list(filter(
        lambda x: getattr(x, ColReportingStatus)=='Pipeline' \
            and getattr(x, ColECSDQuarter)=='FY19Q4',
        fullnewDataBlock
    ))
    printReportData(Q4Qualified, 'Q4Qualified')
    Q4QualifiedSumDict = getSumDict(Q4Qualified)
    print(Q4QualifiedSumDict)

    Q4Unqualified = list(filter(
        lambda x: getattr(x, ColReportingStatus)=='Unqualified' \
            and getattr(x, ColECSDQuarter)=='FY19Q4',
        fullnewDataBlock
    ))
    printReportData(Q4Unqualified, 'Q4Unqualified')
    Q4UnqualifiedSumDict = getSumDict(Q4Unqualified)
    print(Q4UnqualifiedSumDict)

    try:
        wbReport = openpyxl.load_workbook(filename)
    except FileNotFoundError as err:
        print(err)
        sys.exit()
    else:
        wsReport = wbReport.create_sheet('Report')

    wsReport.cell(row=1, column=2, value='AS Net Revenue')
    wsReport.cell(row=1, column=3, value='Wtd AS Net Revenue')
    wsReport.cell(row=1, column=4, value='SI Net Revenue')
    wsReport.cell(row=1, column=5, value='Wtd SI Net Revenue')
    wsReport.cell(row=1, column=6, value='AO Net Revenue')
    wsReport.cell(row=1, column=7, value='Wtd AO Net Revenue')   
    fill = PatternFill("solid", fgColor='D1D1D1')
    for cn in range(1, 8):
        wsReport.cell(row=1, column=cn).fill = fill

    wsReport.cell(row=2, column=1, value='Q2 Pipeline (Qualified)')
    wsReport.cell(row=3, column=1, value='Q2 Pipeline (Unqualified)')
    wsReport.cell(row=4, column=1, value='Q3 Pipeline (Qualified)')
    wsReport.cell(row=5, column=1, value='Q3 Pipeline (Unqualified)')
    wsReport.cell(row=6, column=1, value='Q4 Pipeline (Qualified)')
    wsReport.cell(row=7, column=1, value='Q4 Pipeline (Unqualified)')
    fill = PatternFill("solid", fgColor='ADD8E6')
    for rn in range(2, 8):
        wsReport.cell(row=rn, column=1).fill = fill

    wsReport.cell(
        row=1, 
        column=8, 
        value='Condition'
    )
    wsReport.cell(
        row=2, 
        column=8, 
        value="Reporting Status = 'Pipeline' and ECSD Quarter = 'FY19Q2'"
    )
    wsReport.cell(
        row=3, 
        column=8, 
        value="Reporting Status = 'Unqualified' and ECSD Quarter = 'FY19Q2'"
    )
    wsReport.cell(
        row=4, 
        column=8, 
        value="Reporting Status = 'Pipeline' and ECSD Quarter = 'FY19Q3'"
    )
    wsReport.cell(
        row=5, 
        column=8, 
        value="Reporting Status = 'Unqualified' and ECSD Quarter = 'FY19Q3'"
    )
    wsReport.cell(
        row=6, 
        column=8, 
        value="Reporting Status = 'Pipeline' and ECSD Quarter = 'FY19Q4'"
    )
    wsReport.cell(
        row=7, 
        column=8, 
        value="Reporting Status = 'Unqualified' and ECSD Quarter = 'FY19Q4'"
    )
    fill = PatternFill("solid", fgColor='FFFF00')
    for rn in range(1, 8):
        wsReport.cell(row=rn, column=8).fill = fill
    
    fill = PatternFill("solid", fgColor='71C671')
    for rn in range(2, 8):
        for cn in range(2, 8):
            wsReport.cell(row=rn, column=cn).fill = fill
    
    def writeValue(ws, row, data):
        if len(data) > 0:
            ws.cell(
                row=row, 
                column=2, 
                value=data.get('ASNetRevenue') if data.get('ASNetRevenue') is not None \
                    and isinstance(data.get('ASNetRevenue'), numbers.Real) else 0
            )
            ws.cell(
                row=row, 
                column=3, 
                value=data.get('WtdASNetRevenue') if data.get('WtdASNetRevenue') is not None \
                    and isinstance(data.get('WtdASNetRevenue'), numbers.Real) else 0
            )
            ws.cell(
                row=row, 
                column=4, 
                value=data.get('SINetRevenue') if data.get('SINetRevenue') is not None \
                    and isinstance(data.get('SINetRevenue'), numbers.Real) else 0
            )
            ws.cell(
                row=row, 
                column=5, 
                value=data.get('WtdSINetRevenue') if data.get('WtdSINetRevenue') is not None \
                    and isinstance(data.get('WtdSINetRevenue'), numbers.Real) else 0
            )
            ws.cell(
                row=row, 
                column=6, 
                value=data.get('AONetRevenue') if data.get('AONetRevenue') is not None \
                    and isinstance(data.get('AONetRevenue'), numbers.Real) else 0
            )
            ws.cell(
                row=row, 
                column=7, 
                value=data.get('WtdAONetRevenue') if data.get('WtdAONetRevenue') is not None \
                    and isinstance(data.get('WtdAONetRevenue'), numbers.Real) else 0
            )
        else:
            for cn in range(2, 8):
                ws.cell(row=row, column=cn, value=0)
    
    writeValue(wsReport, 2, Q2QualifiedSumDict)
    writeValue(wsReport, 3, Q2UnqualifiedSumDict)
    writeValue(wsReport, 4, Q3QualifiedSumDict)
    writeValue(wsReport, 5, Q3UnqualifiedSumDict)
    writeValue(wsReport, 6, Q4QualifiedSumDict)
    writeValue(wsReport, 7, Q4UnqualifiedSumDict)

    wbReport.save(filename)

if __name__ == '__main__':
    conf = getConf()
    printConf(conf)
    comparedDataBlock = compare(conf)
    filename = writeBlock(conf, comparedDataBlock)
    newReport(conf, filename)
